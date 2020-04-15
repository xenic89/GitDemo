import openpyxl


class HomePageData:

    test_Homepage_data = [{"FName": "Dinesh", "LName": "Kannan", "Gender": "Male"},
                            {"FName": "Janaranjani", "LName": "Mohan", "Gender": "Female"}]
    @staticmethod
    def getTestData(self):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\acer\\Desktop\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # rows
            if sheet.cell(row=i, column=1).value == "TC2":
                for j in range(2, sheet.max_column + 1):  # columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]